import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ─── BRAND COLORS ────────────────────────────────────────────
NAVY = "1B2A4A"
GREEN = "3DAA5C"
RED = "C0392B"
YELLOW = "F39C12"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F7FA"
MID_GRAY = "DEE2E6"


def _header_style(ws, row, col, value, bg=NAVY, fg=WHITE, bold=True, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color=WHITE),
        right=Side(style="thin", color=WHITE)
    )
    return cell


def _data_style(ws, row, col, value, number_format=None, bg=None, fg="000000", bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", color=fg, bold=bold, size=10)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left",
                               vertical="center")
    if number_format:
        cell.number_format = number_format
    cell.border = Border(bottom=Side(style="hair", color=MID_GRAY))
    return cell


def _set_column_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _add_logo_header(ws, title: str, subtitle: str = ""):
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"DataCrunch — {title}"
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitle or f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub_cell.font = Font(name="Calibri", size=10, color=NAVY, italic=True)
    sub_cell.fill = PatternFill("solid", fgColor="EEF2F7")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws.row_dimensions[2].height = 20


# ─────────────────────────────────────────────────────────────
# FINANCIAL SUMMARY SHEET
# ─────────────────────────────────────────────────────────────
def _build_financial_summary(wb: openpyxl.Workbook, data: dict):
    ws = wb.active
    ws.title = "Financial Summary"
    ws.sheet_view.showGridLines = False

    period = data.get("period", "N/A")
    company = data.get("company_name", "N/A")
    currency = data.get("currency", "EUR")
    fmt = f'#,##0.00 "{currency}"'

    _add_logo_header(ws, "Financial Summary", f"{company} — {period}")

    headers = ["Category", "Item", "Amount", "Validation"]
    cols = ["A", "B", "C", "D"]
    for i, h in enumerate(headers, start=1):
        _header_style(ws, 3, i, h)

    row = 4
    sections = [
        ("Revenue", data.get("revenue", {})),
        ("Expenses", data.get("expenses", {})),
        ("Assets", data.get("assets", {})),
        ("Liabilities", data.get("liabilities", {})),
    ]

    for section_name, section_data in sections:
        if not section_data:
            continue

        items = section_data.get("items", [])
        total_stated = section_data.get("total_stated")
        total_calculated = section_data.get("total_calculated", 0)
        mismatch = section_data.get("mismatch", False)

        # Section header
        ws.merge_cells(f"A{row}:D{row}")
        sec_cell = ws.cell(row=row, column=1, value=section_name.upper())
        sec_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        sec_cell.fill = PatternFill("solid", fgColor=GREEN)
        sec_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22
        row += 1

        for item in items:
            bg = LIGHT_GRAY if row % 2 == 0 else WHITE
            _data_style(ws, row, 1, section_name, bg=bg)
            _data_style(ws, row, 2, item.get("label", ""), bg=bg)
            _data_style(ws, row, 3, item.get("amount", 0), number_format=fmt, bg=bg)
            _data_style(ws, row, 4, "", bg=bg)
            row += 1

        # Total row
        mismatch_text = ""
        total_bg = WHITE
        total_fg = NAVY
        if mismatch and total_stated is not None:
            diff = total_calculated - total_stated
            mismatch_text = f"⚠️ MISMATCH: stated={total_stated:,.0f}, calc={total_calculated:,.0f}, diff={diff:+,.0f}"
            total_bg = "FDECEA"
            total_fg = RED
        elif total_stated is not None:
            mismatch_text = "✅ Validated"

        _data_style(ws, row, 1, "", bold=True, bg="E8F5E9")
        _data_style(ws, row, 2, f"TOTAL {section_name}", bold=True, bg="E8F5E9", fg=NAVY)
        _data_style(ws, row, 3, total_calculated, number_format=fmt, bold=True, bg="E8F5E9", fg=NAVY)
        _data_style(ws, row, 4, mismatch_text, bg=total_bg, fg=total_fg, bold=mismatch)
        ws.row_dimensions[row].height = 20
        row += 2

    # Key metrics
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    kpi_cell = ws.cell(row=row, column=1, value="KEY METRICS")
    kpi_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    kpi_cell.fill = PatternFill("solid", fgColor=NAVY)
    kpi_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    row += 1

    for label, value in [("EBITDA", data.get("ebitda")), ("Net Income", data.get("net_income"))]:
        if value is not None:
            _data_style(ws, row, 1, "KPI")
            _data_style(ws, row, 2, label)
            _data_style(ws, row, 3, value, number_format=fmt, bold=True, fg=NAVY)
            _data_style(ws, row, 4, "")
            row += 1

    _set_column_widths(ws, {"A": 18, "B": 35, "C": 20, "D": 50})


# ─────────────────────────────────────────────────────────────
# REVENUE PER CLIENT SHEET
# ─────────────────────────────────────────────────────────────
def _build_revenue_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Revenue per Client")
    ws.sheet_view.showGridLines = False

    currency = data.get("currency", "EUR")
    fmt = f'#,##0.00 "{currency}"'
    pct_fmt = '0.0"%"'

    _add_logo_header(ws, "Revenue per Client",
                     f"{data.get('company_name', '')} — {data.get('period', '')}")

    headers = ["#", "Client Name", "Revenue", "% of Total", "Contract Type", "Flag"]
    for i, h in enumerate(headers, start=1):
        _header_style(ws, 3, i, h)

    clients = data.get("clients", [])
    total_calculated = data.get("total_calculated", 0)
    total_stated = data.get("total_stated")
    mismatch = data.get("mismatch", False)

    for idx, client in enumerate(clients, start=1):
        row = idx + 3
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        revenue = client.get("revenue", 0)
        pct = (revenue / total_calculated * 100) if total_calculated else 0

        _data_style(ws, row, 1, idx, bg=bg)
        _data_style(ws, row, 2, client.get("name", ""), bg=bg)
        _data_style(ws, row, 3, revenue, number_format=fmt, bg=bg)
        _data_style(ws, row, 4, pct, number_format=pct_fmt, bg=bg)
        _data_style(ws, row, 5, client.get("contract_type", ""), bg=bg)

        # Flag high concentration (>30%)
        flag = "⚠️ High concentration" if pct > 30 else ""
        flag_fg = RED if pct > 30 else "000000"
        _data_style(ws, row, 6, flag, fg=flag_fg, bg=bg)

    # Totals row
    total_row = len(clients) + 4
    validation_text = ""
    if mismatch and total_stated:
        diff = total_calculated - total_stated
        validation_text = f"⚠️ MISMATCH: diff={diff:+,.0f}"
        total_bg = "FDECEA"
    else:
        validation_text = "✅ Validated" if total_stated else ""
        total_bg = "E8F5E9"

    ws.merge_cells(f"A{total_row}:B{total_row}")
    _data_style(ws, total_row, 1, "TOTAL", bold=True, bg=total_bg, fg=NAVY)
    _data_style(ws, total_row, 3, total_calculated, number_format=fmt, bold=True, bg=total_bg, fg=NAVY)
    _data_style(ws, total_row, 4, 100.0, number_format=pct_fmt, bold=True, bg=total_bg)
    _data_style(ws, total_row, 6, validation_text, bold=True,
                fg=RED if mismatch else GREEN, bg=total_bg)

    _set_column_widths(ws, {"A": 6, "B": 35, "C": 20, "D": 15, "E": 20, "F": 35})


# ─────────────────────────────────────────────────────────────
# PAYROLL SHEET
# ─────────────────────────────────────────────────────────────
def _build_payroll_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Payroll")
    ws.sheet_view.showGridLines = False

    currency = data.get("currency", "EUR")
    fmt = f'#,##0.00 "{currency}"'

    _add_logo_header(ws, "Payroll Analysis",
                     f"{data.get('company_name', '')} — {data.get('period', '')}")

    headers = ["#", "Employee", "Role", "Department", "Gross Salary", "Net Salary", "Social Charges"]
    for i, h in enumerate(headers, start=1):
        _header_style(ws, 3, i, h)

    employees = data.get("employees", [])
    total_gross_calculated = data.get("total_gross_calculated", 0)
    total_gross_stated = data.get("total_gross_stated")
    mismatch = data.get("mismatch", False)

    for idx, emp in enumerate(employees, start=1):
        row = idx + 3
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        _data_style(ws, row, 1, idx, bg=bg)
        _data_style(ws, row, 2, emp.get("name", ""), bg=bg)
        _data_style(ws, row, 3, emp.get("role", ""), bg=bg)
        _data_style(ws, row, 4, emp.get("department", ""), bg=bg)
        _data_style(ws, row, 5, emp.get("gross_salary", 0), number_format=fmt, bg=bg)
        _data_style(ws, row, 6, emp.get("net_salary"), number_format=fmt, bg=bg)
        _data_style(ws, row, 7, emp.get("social_charges"), number_format=fmt, bg=bg)

    # Totals row
    total_row = len(employees) + 4
    validation_text = ""
    if mismatch and total_gross_stated:
        diff = total_gross_calculated - total_gross_stated
        validation_text = f"⚠️ MISMATCH: diff={diff:+,.0f}"
        total_bg = "FDECEA"
    else:
        validation_text = f"✅ Validated | Headcount: {data.get('headcount', len(employees))}"
        total_bg = "E8F5E9"

    ws.merge_cells(f"A{total_row}:D{total_row}")
    _data_style(ws, total_row, 1, f"TOTAL — {data.get('headcount', len(employees))} employees",
                bold=True, bg=total_bg, fg=NAVY)
    _data_style(ws, total_row, 5, total_gross_calculated, number_format=fmt, bold=True, bg=total_bg, fg=NAVY)
    _data_style(ws, total_row, 6, data.get("total_net_calculated"), number_format=fmt, bold=True, bg=total_bg)
    _data_style(ws, total_row, 7, validation_text, bold=mismatch,
                fg=RED if mismatch else GREEN, bg=total_bg)

    _set_column_widths(ws, {"A": 6, "B": 30, "C": 25, "D": 20, "E": 18, "F": 18, "G": 35})


# ─────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────
def generate_excel(extracted_data: dict, output_path: str) -> str:
    """
    Generates a professional Excel file from extracted AI data.
    Returns the output file path.
    """
    wb = openpyxl.Workbook()
    doc_type = extracted_data.get("document_type", "financial_statement")

    if doc_type == "financial_statement":
        _build_financial_summary(wb, extracted_data)
    elif doc_type == "revenue_list":
        # For revenue list, use first sheet as revenue sheet
        ws = wb.active
        ws.title = "Revenue per Client"
        wb.remove(ws)
        _build_revenue_sheet(wb, extracted_data)
    elif doc_type == "payroll":
        ws = wb.active
        ws.title = "Payroll"
        wb.remove(ws)
        _build_payroll_sheet(wb, extracted_data)
    else:
        _build_financial_summary(wb, extracted_data)

    # Add validation summary sheet
    _build_validation_sheet(wb, extracted_data)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def _build_validation_sheet(wb: openpyxl.Workbook, data: dict):
    ws = wb.create_sheet("Validation Report")
    ws.sheet_view.showGridLines = False

    _add_logo_header(ws, "Validation Report", "AI Extraction Quality Check")

    _header_style(ws, 3, 1, "Check")
    _header_style(ws, 3, 2, "Status")
    _header_style(ws, 3, 3, "Details")

    checks = []

    def add_check(section_data, name):
        if not section_data:
            return
        mismatch = section_data.get("mismatch", False)
        stated = section_data.get("total_stated") or section_data.get("total_gross_stated")
        calculated = section_data.get("total_calculated") or section_data.get("total_gross_calculated", 0)
        notes = section_data.get("validation_notes", "")

        if stated is not None:
            status = "⚠️ MISMATCH" if mismatch else "✅ OK"
            detail = f"Stated: {stated:,.0f} | Calculated: {calculated:,.0f}" + (
                f" | Diff: {calculated - stated:+,.0f}" if mismatch else ""
            )
            checks.append((name, status, detail + (f" | {notes}" if notes else "")))

    for section in ["revenue", "expenses", "assets", "liabilities"]:
        add_check(data.get(section, {}), section.title())

    if data.get("document_type") in ("revenue_list", "payroll"):
        add_check(data, data.get("document_type", "").replace("_", " ").title())

    for i, (check, status, detail) in enumerate(checks, start=4):
        bg = "FDECEA" if "MISMATCH" in status else "E8F5E9"
        fg = RED if "MISMATCH" in status else GREEN
        _data_style(ws, i, 1, check, bg=bg)
        _data_style(ws, i, 2, status, fg=fg, bold=True, bg=bg)
        _data_style(ws, i, 3, detail, bg=bg)

    notes = data.get("validation_notes", "")
    if notes:
        row = len(checks) + 5
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1, value=f"AI Notes: {notes}").font = Font(
            italic=True, color=NAVY, size=9
        )

    _set_column_widths(ws, {"A": 20, "B": 18, "C": 70})
